import time
import sys
import requests
from bs4 import BeautifulSoup
import openpyxl
import re

def loading_animation():
    animation = "|/-\\"
    idx = 0
    while True:
        print(f"Loading... {animation[idx % len(animation)]}", end="\r")
        idx += 1
        time.sleep(0.2)  # Увеличили время ожидания
        if idx == len(animation) * 5:  # Увеличили количество итераций
            break

def color_conditioner(lst_url):
    conditioner_data = []
    lst_name_model = None
    url = 'https://catalog.onliner.by/conditioners/gree/'

    for key, value in lst_url.items():
        for model in value:
            try:
                response = requests.get(url + model)
                response.raise_for_status()  # Проверяем статус ответа
                soup = BeautifulSoup(response.content, 'html.parser')
                atribut_a = soup.find_all('a', class_='offers-description-filter-control offers-description-filter-control_switcher js-facet-configurations-link')
                lst_name_model = value + ([x['href'] for x in atribut_a])
            except requests.exceptions.RequestException as e:
                print(f"Error fetching data for model {model}: {e}")
                continue  # Пропускаем модель при ошибке

        for data in lst_name_model:
            try:
                re_data = re.search(r'[^/]*$', data).group()
                json_url = f'https://catalog.onliner.by/sdapi/shop.api/products/{re_data}/positions?town_id=17030&limit_total=6&has_delivery=1'
                get_json_data = requests.get(json_url)
                get_json_data.raise_for_status()  # Проверяем статус ответа
                json_html = get_json_data.json()
                conditioner_data.append({
                    'Model': re_data,
                    'Shop': []
                })
                id_shop = [name_shop for name_shop in json_html['shops']]
                price = [price_shop['position_price'] for price_shop in json_html['positions']['primary']]

                for title, price in zip(id_shop, price):
                    conditioner_data[-1]['Shop'].append({
                        'Shop': json_html['shops'][title]['title'],
                        'Price': f"{price['amount']} {price['currency']}"
                    })
            except (requests.exceptions.RequestException, KeyError) as e:
                print(f"Error processing data for model {re_data}: {e}")
                continue  # Пропускаем модель при ошибке

    return conditioner_data

if __name__ == "__main__":
    try:
        loading_animation()
        lst_url = {
            'pular': ['gwh09agaxak6dna4'],
            'bora': ['gwh09aaaxak6dna2'],
            'arctic': ['gwh09qcxbk6dnc2f'],
            'airy': ['gwh09avcxbk6dnaw', 'gwh09avcxbk6dnab', 'gwh09avcxbk6dnac'],
            'lyra': ['gwh09acck6dna1fw', 'gwh09acck6dna1f', 'gwh09acck6dna1fh']
        }
        result = color_conditioner(lst_url)

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = 'Модель'
        sheet['B1'] = 'Название магазина'
        sheet['C1'] = 'Цена'

        row = 2
        for conditioner in result:
            sheet.cell(row=row, column=1, value=(conditioner['Model']).upper())
            for shop in conditioner['Shop']:
                sheet.cell(row=row, column=2, value=shop['Shop'])
                sheet.cell(row=row, column=3, value=shop['Price'])
                row += 1
            row += 1

        workbook.save('conditioner_data.xlsx')
        print("Загрузка завершена!")
    except (requests.exceptions.RequestException, KeyError, Exception) as e:
        print(f"Ошибка: {e}")
        sys.exit(1)
